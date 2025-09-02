import json
import re
from dataclasses import dataclass
from pathlib import Path
from typing import Dict, List, Optional, Tuple, Union


@dataclass
class ApiEntry:
    api_id: str
    url: str
    method: Optional[str]
    sheet: str
    row: int


def find_excel_file(explicit: Optional[str] = None) -> Path:
    if explicit:
        p = Path(explicit)
        if not p.exists():
            raise FileNotFoundError(f"Excel file not found: {explicit}")
        return p
    # Prefer filenames that look like Kiwoom REST API doc
    candidates: List[Path] = list(Path('.').glob('*.xlsx'))
    scored: List[Tuple[int, Path]] = []
    for c in candidates:
        name = c.name.lower()
        score = 0
        if 'rest' in name:
            score += 2
        if 'api' in name:
            score += 2
        if 'kiwoom' in name or '키움' in name or 'Ű' in name:  # handle mojibake
            score += 2
        scored.append((score, c))
    if not scored:
        raise FileNotFoundError("No .xlsx files found in current directory")
    scored.sort(key=lambda x: x[0], reverse=True)
    best = scored[0][1]
    return best


def normalize(s: Optional[Union[str, int, float]]) -> str:
    if s is None:
        return ''
    if isinstance(s, (int, float)):
        s = str(s)
    s = str(s).strip()
    return s


def parse_workbook(xlsx_path: Path) -> List[ApiEntry]:
    try:
        from openpyxl import load_workbook  # type: ignore
    except Exception as e:  # pragma: no cover
        raise RuntimeError("Requires openpyxl. Install with: pip install openpyxl") from e

    wb = load_workbook(filename=str(xlsx_path), data_only=True, read_only=True)
    results: List[ApiEntry] = []
    # header detectors
    api_id_keys = {"api id", "apiid", "api_id"}
    url_keys = {"url"}
    method_keys = {"method", "http method"}

    for ws in wb.worksheets:
        # Scan first 50 rows for header row containing API ID and URL
        header_row_idx = None
        header_map: Dict[str, int] = {}
        for r_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=50, values_only=True), start=1):
            # build normalized header names
            col_map: Dict[str, int] = {}
            for c_idx, val in enumerate(row, start=1):
                label = normalize(val).lower()
                if not label:
                    continue
                key = re.sub(r"\s+", " ", label)
                if key in api_id_keys:
                    col_map['api_id'] = c_idx
                if key in url_keys:
                    col_map['url'] = c_idx
                if key in method_keys:
                    col_map['method'] = c_idx
            if 'api_id' in col_map and 'url' in col_map:
                header_row_idx = r_idx
                header_map = col_map
                break
        if header_row_idx is None:
            # Try a looser heuristic: any row with URL and a cell that matches an api id pattern like [a-z]{2}[0-9]{5} or [0-9A-Z]{2}
            apiid_pat = re.compile(r"^[A-Za-z]{2}\d{5}$|^[0-9A-Za-z]{2}$")
            for r_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=50, values_only=True), start=1):
                url_col = None
                api_col = None
                for c_idx, val in enumerate(row, start=1):
                    v = normalize(val)
                    if v.upper() == 'URL':
                        url_col = c_idx
                    if v.upper() in ('API ID', 'APIID', 'API_ID'):
                        api_col = c_idx
                if url_col and api_col:
                    header_row_idx = r_idx
                    header_map = {'api_id': api_col, 'url': url_col}
                    break
        if header_row_idx is None:
            continue

        # Iterate rows until a long blank streak
        blank_streak = 0
        for r_idx, row in enumerate(ws.iter_rows(min_row=header_row_idx + 1, values_only=True), start=header_row_idx + 1):
            api_id = normalize(row[header_map['api_id'] - 1]) if header_map.get('api_id') else ''
            url = normalize(row[header_map['url'] - 1]) if header_map.get('url') else ''
            method = normalize(row[header_map['method'] - 1]) if header_map.get('method') else None
            if not api_id and not url:
                blank_streak += 1
                if blank_streak >= 10:
                    break
                continue
            blank_streak = 0
            # Basic validation
            if not api_id or not url:
                continue
            # Keep URL path part only if a full URL
            if url.startswith('http://') or url.startswith('https://'):
                try:
                    # crude extraction of path
                    url = '/' + url.split('://', 1)[1].split('/', 1)[1]
                except Exception:
                    pass
            results.append(ApiEntry(api_id=api_id, url=url, method=method if method else None, sheet=ws.title, row=r_idx))

    return results


def build_outputs(entries: List[ApiEntry]) -> Tuple[Dict[str, str], Dict[str, Union[Dict[str, Union[str, int]], List[Dict[str, Union[str, int]]]]]]:
    simple: Dict[str, str] = {}
    full: Dict[str, Union[Dict[str, Union[str, int]], List[Dict[str, Union[str, int]]]]] = {}
    for e in entries:
        # simple mapping (first occurrence wins)
        simple.setdefault(e.api_id, e.url)
        detail = {"url": e.url, "method": e.method or '', "sheet": e.sheet, "row": e.row}
        if e.api_id in full:
            curr = full[e.api_id]
            if isinstance(curr, list):
                curr.append(detail)
            else:
                full[e.api_id] = [curr, detail]  # convert to list to preserve all
        else:
            full[e.api_id] = detail
    return simple, full


def main() -> None:
    # Resolve excel file
    excel_file = None
    # Try exact Korean name first
    exact = Path('키움 REST API 문서.xlsx')
    if exact.exists():
        excel_file = exact
    else:
        excel_file = find_excel_file()
    entries = parse_workbook(excel_file)
    simple, full = build_outputs(entries)

    Path('restapi_map_from_excel.json').write_text(json.dumps(simple, ensure_ascii=False, indent=2), encoding='utf-8')
    Path('restapi_full_from_excel.json').write_text(json.dumps(full, ensure_ascii=False, indent=2), encoding='utf-8')
    print(f"Parsed {len(entries)} entries from {excel_file.name} across all sheets")
    print(f"Wrote restapi_map_from_excel.json ({len(simple)} unique API IDs)")
    print(f"Wrote restapi_full_from_excel.json")


if __name__ == '__main__':
    main()

