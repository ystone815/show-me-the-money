import json
import re
from dataclasses import dataclass, field
from pathlib import Path
from typing import Dict, List, Optional, Tuple, Any


@dataclass
class FieldRow:
    element: str = ""
    name_kr: str = ""
    ftype: str = ""
    require: str = ""
    length: str = ""
    desc: str = ""


@dataclass
class ApiSpec:
    api_id: str
    url: str = ""
    method: str = ""
    sheet: str = ""
    row: int = 0
    request_headers: List[FieldRow] = field(default_factory=list)
    request_body: List[FieldRow] = field(default_factory=list)
    response_headers: List[FieldRow] = field(default_factory=list)
    response_body: List[FieldRow] = field(default_factory=list)
    request_example: str = ""
    response_example: str = ""


def normalize(s: Any) -> str:
    if s is None:
        return ""
    if isinstance(s, (int, float)):
        s = str(s)
    s = str(s)
    return s.strip()


def normkey(s: Any) -> str:
    s = normalize(s).lower()
    s = re.sub(r"\s+", " ", s)
    return s


def find_api_blocks(ws) -> List[Tuple[int, str]]:
    """Return list of tuples (row_index, api_id) indicating API blocks."""
    blocks: List[Tuple[int, str]] = []
    pat = re.compile(r"^[A-Za-z]{2}\d{5}$|^[0-9A-Za-z]{2}$")
    max_row = ws.max_row or 0
    max_col = ws.max_column or 0
    for r in range(1, max_row + 1):
        for c in range(1, max_col + 1):
            v = ws.cell(row=r, column=c).value
            if normkey(v) in ("api id", "apiid", "api_id"):
                # Look to the right up to +3 cols for ID
                for dc in range(1, 4):
                    v2 = ws.cell(row=r, column=c + dc).value
                    s2 = normalize(v2)
                    if pat.match(s2 or ""):
                        blocks.append((r, s2))
                        break
    # Deduplicate by (row, id)
    dedup: List[Tuple[int, str]] = []
    seen = set()
    for r, aid in sorted(blocks, key=lambda x: x[0]):
        k = (r, aid)
        if k in seen:
            continue
        seen.add(k)
        dedup.append((r, aid))
    return dedup


def find_nearby_label(ws, start_row: int, label: str, search_rows: int = 60) -> Optional[Tuple[int, int]]:
    """Find a cell with text == label (case-insensitive) within rows below start_row.
    Returns (row, col) or None.
    """
    label_l = label.lower()
    max_row = ws.max_row or 0
    max_col = ws.max_column or 0
    end_row = min(max_row, start_row + search_rows)
    for r in range(start_row, end_row + 1):
        for c in range(1, max_col + 1):
            if normkey(ws.cell(row=r, column=c).value) == label_l:
                return (r, c)
    return None


def find_url_and_method(ws, start_row: int) -> Tuple[str, str]:
    url = ""
    method = ""
    pos = find_nearby_label(ws, start_row, "url", search_rows=80)
    if pos:
        r, c = pos
        url = normalize(ws.cell(row=r, column=c + 1).value)
        if url.startswith("http://") or url.startswith("https://"):
            try:
                url = "/" + url.split("://", 1)[1].split("/", 1)[1]
            except Exception:
                pass

    pos_m = find_nearby_label(ws, start_row, "method", search_rows=80)
    if pos_m:
        r, c = pos_m
        method = normalize(ws.cell(row=r, column=c + 1).value)
    return url, method


def detect_table_header(row_vals: List[str]) -> Dict[str, int]:
    # Accept many variants
    header_aliases = {
        "element": {"element", "영문", "항목", "key"},
        "name_kr": {"한글", "한글명", "설명", "이름"},
        "ftype": {"type", "자료형", "타입"},
        "require": {"require", "required", "필수"},
        "length": {"length", "len", "길이"},
        "desc": {"description", "desc", "비고", "설명"},
    }
    idx: Dict[str, int] = {}
    for i, v in enumerate(row_vals):
        k = normkey(v)
        for key, aliases in header_aliases.items():
            if k in aliases:
                idx[key] = i
    # Need at least element + type to consider a table
    if "element" in idx and ("ftype" in idx or "desc" in idx or "require" in idx):
        return idx
    return {}


def parse_field_table(ws, start_row: int, max_rows: int = 200) -> Tuple[List[FieldRow], int]:
    """Parse a table starting at or after start_row where the first non-empty row is header.
    Returns (rows, next_row) where next_row is the row after the table.
    """
    max_row = ws.max_row or 0
    end_row = min(max_row, start_row + max_rows)
    header_idx: Dict[str, int] = {}
    header_row = None
    # find header row within next 10 rows
    for r in range(start_row, min(end_row, start_row + 10) + 1):
        vals = [normalize(ws.cell(row=r, column=c).value) for c in range(1, ws.max_column + 1)]
        if any(vals):
            header_idx = detect_table_header(vals)
            if header_idx:
                header_row = r
                break
    rows: List[FieldRow] = []
    if not header_idx or header_row is None:
        return rows, start_row
    # parse until blank streak
    blank_streak = 0
    for r in range(header_row + 1, end_row + 1):
        vals = [normalize(ws.cell(row=r, column=c).value) for c in range(1, ws.max_column + 1)]
        if not any(vals):
            blank_streak += 1
            if blank_streak >= 3:
                return rows, r
            continue
        blank_streak = 0
        fr = FieldRow(
            element=vals[header_idx.get("element", -1)] if header_idx.get("element") is not None else "",
            name_kr=vals[header_idx.get("name_kr", -1)] if header_idx.get("name_kr") is not None else "",
            ftype=vals[header_idx.get("ftype", -1)] if header_idx.get("ftype") is not None else "",
            require=vals[header_idx.get("require", -1)] if header_idx.get("require") is not None else "",
            length=vals[header_idx.get("length", -1)] if header_idx.get("length") is not None else "",
            desc=vals[header_idx.get("desc", -1)] if header_idx.get("desc") is not None else "",
        )
        # skip garbage rows lacking element and desc
        if not fr.element and not fr.desc:
            continue
        rows.append(fr)
    return rows, end_row


def collect_example(ws, label_pos: Tuple[int, int], until_labels: List[str]) -> str:
    r0, _ = label_pos
    max_row = ws.max_row or 0
    lines: List[str] = []
    until_set = {l.lower() for l in until_labels}
    for r in range(r0 + 1, min(max_row, r0 + 200) + 1):
        row_vals = [ws.cell(row=r, column=c).value for c in range(1, ws.max_column + 1)]
        # stop if next section label found
        if any(normkey(v) in until_set for v in row_vals):
            break
        # accumulate non-empty cells joined by tabs
        text = "\t".join([normalize(v) for v in row_vals if normalize(v)])
        if text:
            lines.append(text)
    # Heuristic: join and trim
    return "\n".join(lines).strip()


def parse_api_spec(ws, start_row: int, api_id: str) -> ApiSpec:
    url, method = find_url_and_method(ws, start_row)
    spec = ApiSpec(api_id=api_id, url=url, method=method or "", sheet=ws.title, row=start_row)

    # Find Request section
    req_pos = find_nearby_label(ws, start_row, "request", search_rows=200)
    if req_pos:
        # optional sub markers 'Header' and 'Body'
        # Header table
        hdr_pos = find_nearby_label(ws, req_pos[0], "header", search_rows=10)
        parse_start = (hdr_pos[0] + 1) if hdr_pos else (req_pos[0] + 1)
        rows, nxt = parse_field_table(ws, parse_start)
        spec.request_headers = rows
        # Body table
        body_pos = find_nearby_label(ws, nxt, "body", search_rows=20)
        if not body_pos:
            # maybe body appears directly after request when header omitted
            body_pos = find_nearby_label(ws, req_pos[0], "body", search_rows=60)
        if body_pos:
            body_rows, _ = parse_field_table(ws, body_pos[0] + 1)
            spec.request_body = body_rows

    # Find Response section
    res_pos = find_nearby_label(ws, start_row, "response", search_rows=260)
    if res_pos:
        hdr_pos = find_nearby_label(ws, res_pos[0], "header", search_rows=10)
        parse_start = (hdr_pos[0] + 1) if hdr_pos else (res_pos[0] + 1)
        rows, nxt = parse_field_table(ws, parse_start)
        spec.response_headers = rows
        body_pos = find_nearby_label(ws, nxt, "body", search_rows=20)
        if not body_pos:
            body_pos = find_nearby_label(ws, res_pos[0], "body", search_rows=60)
        if body_pos:
            body_rows, _ = parse_field_table(ws, body_pos[0] + 1)
            spec.response_body = body_rows

    # Examples
    req_ex_pos = find_nearby_label(ws, start_row, "request example", search_rows=300)
    if req_ex_pos:
        spec.request_example = collect_example(ws, req_ex_pos, until_labels=["response example", "response", "api id", "url"])
    res_ex_pos = find_nearby_label(ws, start_row, "response example", search_rows=300)
    if res_ex_pos:
        spec.response_example = collect_example(ws, res_ex_pos, until_labels=["request example", "request", "api id", "url"])

    return spec


def main() -> None:
    try:
        from openpyxl import load_workbook  # type: ignore
    except Exception as e:  # pragma: no cover
        raise SystemExit("Requires openpyxl. Install with: pip install openpyxl")

    # Resolve excel file by preference
    excel = None
    for cand in ["키움 REST API 문서.xlsx", "Ű�� REST API ����.xlsx"]:
        p = Path(cand)
        if p.exists():
            excel = p
            break
    if excel is None:
        xs = list(Path('.').glob('*.xlsx'))
        if not xs:
            raise SystemExit("No .xlsx found")
        excel = xs[0]

    wb = load_workbook(filename=str(excel), data_only=True, read_only=True)

    specs: Dict[str, Dict[str, Any]] = {}
    for ws in wb.worksheets:
        blocks = find_api_blocks(ws)
        for start_row, api_id in blocks:
            spec = parse_api_spec(ws, start_row, api_id)
            # Merge with existing if duplicate api_id: prefer first non-empty url/method and extend fields
            if api_id not in specs:
                d = spec.__dict__.copy()
                # dataclasses contain lists of FieldRow objects; convert to dicts
                for k in ["request_headers", "request_body", "response_headers", "response_body"]:
                    d[k] = [fr.__dict__ for fr in getattr(spec, k)]
                specs[api_id] = d
            else:
                prev = specs[api_id]
                if not prev.get("url") and spec.url:
                    prev["url"] = spec.url
                if not prev.get("method") and spec.method:
                    prev["method"] = spec.method
                # extend lists if new items found
                for k in ["request_headers", "request_body", "response_headers", "response_body"]:
                    prev_list = prev.get(k, [])
                    new_list = [fr.__dict__ for fr in getattr(spec, k)]
                    if new_list:
                        prev_list.extend(new_list)
                        prev[k] = prev_list
                for k in ["request_example", "response_example"]:
                    if not prev.get(k) and getattr(spec, k):
                        prev[k] = getattr(spec, k)
                specs[api_id] = prev

    # Dump outputs
    out = Path("restapi_specs_from_excel.json")
    out.write_text(json.dumps(specs, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"Parsed specs for {len(specs)} API IDs from {excel.name}")
    print(f"Wrote {out}")


if __name__ == "__main__":
    main()

